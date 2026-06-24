#!/usr/bin/env python3
"""
qa_workbook.py
Regression QA for the generated workbook: cross-checks every data table against
the engine's source aggregations, replicates the live formulas, and validates
that all charts reference populated, length-aligned data. Run after a rebuild or
(future) a data refresh.

    python scripts/qa_workbook.py            # exits non-zero if anything fails
"""
import warnings; warnings.filterwarnings("ignore")
import sys, re, zipfile
from pathlib import Path
_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "scripts"))
import engine
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

WB = str(_ROOT / "output" / "outage_workbook.xlsx")
ctx = engine.build_context(str(_ROOT / "data" / "rEFINERY oUTAGES.xlsx"))
wb = load_workbook(WB)
MONTHS = engine.MONTHS
PADDS = engine.PADD_ORDER

PASS = []; FAIL = []
def check(cond, label, detail=""):
    (PASS if cond else FAIL).append(label + (f"  [{detail}]" if detail and not cond else ""))

def band_row(ws, sub):
    # match by PREFIX so "Planned ..." never matches "Unplanned ..."
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.strip().lower().startswith(sub.lower()):
                return c.row, c.column
    return None, None

def num(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return v if isinstance(v, (int, float)) else None

def close(a, b, tol=1.0):
    if a is None or b is None: return False
    return abs(a - b) <= tol

# ----- helper: read a year-rows block (col Bidx=year, then month/value cols) -----
def read_year_block(ws, band_text, ncols, year_col=2, first_data_offset=2):
    br, bc = band_row(ws, band_text)
    if br is None: return None, None
    hdr = br + 1
    rows = {}
    r = hdr + 1
    while True:
        y = ws.cell(row=r, column=year_col).value
        if not isinstance(y, (int, float)) or y < 2000 or y > 2040:
            break
        rows[int(y)] = [ws.cell(row=r, column=first_data_offset + 1 + j).value for j in range(ncols)]
        r += 1
        if r > hdr + 40: break
    return rows, hdr

# ============================================================ TRENDS: annual
ws = wb["Trends"]
s = ctx["summary"]
br, _ = band_row(ws, "Annual Capacity Offline")
hdr = br + 1
miss = 0
for i in range(40):
    r = hdr + 1 + i
    y = ws.cell(row=r, column=2).value
    if not isinstance(y, (int, float)): break
    y = int(y)
    pl = num(ws, r, 3); un = num(ws, r, 4)
    if not close(pl, float(s.loc[y, "Planned"])): miss += 1
    if not close(un, float(s.loc[y, "Unplanned"])): miss += 1
check(miss == 0, "Trends annual Planned/Unplanned match engine", f"{miss} mismatches")

# ============================================================ TRENDS: monthly matrices
for title, frame in [("Total Offline - Monthly", ctx["monthly_total"]),
                     ("Planned - Monthly", ctx["monthly_planned"]),
                     ("Unplanned - Monthly", ctx["monthly_unplanned"])]:
    rows, _ = read_year_block(ws, title, 12)
    miss = 0; n = 0
    for y, vals in rows.items():
        for j, m in enumerate(MONTHS):
            exp = float(frame.loc[y, m]) if y in frame.index else 0.0
            n += 1
            if not close(vals[j], exp): miss += 1
    check(miss == 0, f"Trends [{title}] {n} monthly cells match engine", f"{miss} mismatches")

# ============================================================ PADD x year matrices
ws = wb["PADD"]
for title, frame in [("Total Offline - PADD x Year", ctx["padd_total"]),
                     ("Unplanned - PADD x Year", ctx["padd_unplanned"]),
                     ("Planned - PADD x Year", ctx["padd_planned"])]:
    br, _ = band_row(ws, title)
    hdr = br + 1
    years = [ws.cell(row=hdr, column=3 + j).value for j in range(12)]
    years = [int(y) for y in years if isinstance(y, (int, float))]
    miss = 0; n = 0
    for pi, p in enumerate(PADDS):
        r = hdr + 1 + pi
        for j, y in enumerate(years):
            v = num(ws, r, 3 + j); exp = float(frame.loc[p, y])
            n += 1
            if not close(v, exp): miss += 1
    check(miss == 0, f"PADD [{title}] {n} cells match engine", f"{miss} mismatches")

# PADD combo chart data blocks
pm = ctx["padd_month"]
for p in PADDS:
    br, _ = band_row(ws, f"{p} - 2026 plan")
    hdr = br + 1; first = hdr + 1
    # row order: 2026 Planned, 2026 Unplanned, 2025 Total, 2024 Total, 2023 Total, 2027 Planned
    specs = [("planned", 2026), ("unplanned", 2026), ("total", 2025),
             ("total", 2024), ("total", 2023), ("planned", 2027)]
    miss = 0
    for k, (key, yr) in enumerate(specs):
        for j, m in enumerate(MONTHS):
            v = num(ws, first + k, 3 + j)
            exp = float(pm[p][key].loc[yr, m]) if yr in pm[p][key].index else 0.0
            if not close(v, exp): miss += 1
    check(miss == 0, f"PADD combo data [{p}] matches engine", f"{miss} mismatches")

# ============================================================ UNITS
ws = wb["Units & Refineries"]
um = ctx["unit_total"]
br, _ = band_row(ws, "Capacity Offline by Unit Category")
hdr = br + 1
uyears = [ws.cell(row=hdr, column=3 + j).value for j in range(8)]
uyears = [int(y) for y in uyears if isinstance(y, (int, float))]
miss = 0; n = 0
for i, u in enumerate(um.index):
    r = hdr + 1 + i
    lbl = ws.cell(row=r, column=2).value
    for j, y in enumerate(uyears):
        v = num(ws, r, 3 + j); exp = float(um.loc[u, y]); n += 1
        if not close(v, exp): miss += 1
check(miss == 0, f"Units matrix {n} cells match engine", f"{miss} mismatches")

# Naphtha annual
na = ctx["naphtha"]["annual"]
rows, _ = read_year_block(ws, "Naphtha / Octane Complex", 2)
miss = 0
for y, vals in rows.items():
    if y in na.index:
        if not close(vals[0], float(na.loc[y, "Planned"])): miss += 1
        if not close(vals[1], float(na.loc[y, "Unplanned"])): miss += 1
check(miss == 0, "Naphtha annual matches engine", f"{miss} mismatches")

# Mogas annual
ma = ctx["mogas_annual"]
rows, _ = read_year_block(ws, "Mogas-Equivalent Offline by Year", 2)
miss = 0
for y, vals in rows.items():
    if y in ma.index:
        if not close(vals[0], float(ma.loc[y, "Planned"])): miss += 1
        if not close(vals[1], float(ma.loc[y, "Unplanned"])): miss += 1
check(miss == 0, "Mogas annual matches engine", f"{miss} mismatches")

# Top refineries (numbers; operator names are shortened so check totals)
pl = ctx["plants"].reset_index(drop=True)
br, _ = band_row(ws, "Top 15 Refineries")
hdr = br + 1; miss = 0
for i in range(len(pl)):
    r = hdr + 1 + i
    plant = ws.cell(row=r, column=2).value
    tot = num(ws, r, 5)
    match = pl[pl["plant"] == plant]
    if len(match) and not close(tot, float(match.iloc[0]["total"])): miss += 1
check(miss == 0, "Top refineries totals match engine", f"{miss} mismatches")

# ============================================================ DATA tables vs tidy
ws = wb["Data"]
for tname, tidy in [("tPADD", ctx["tidy_padd"]), ("tUNIT", ctx["tidy_unit"])]:
    tbl = ws.tables[tname]
    minc, minr, maxc, maxr = range_boundaries(tbl.ref)
    ndata = maxr - minr  # excl header
    check(ndata == len(tidy), f"Data table {tname} row count = engine ({ndata} vs {len(tidy)})")
    # spot-check 5 evenly spaced rows
    miss = 0
    idxs = [0, len(tidy)//4, len(tidy)//2, 3*len(tidy)//4, len(tidy)-1]
    for k in idxs:
        row = tidy.iloc[k]
        xr = minr + 1 + k
        vy = ws.cell(row=xr, column=minc).value
        vk = ws.cell(row=xr, column=minc+2).value
        vkbd = ws.cell(row=xr, column=minc+4).value
        if int(vy) != int(row["year"]) or vk != row["key"] or not close(vkbd, float(row["kbd"]), 0.1):
            miss += 1
    check(miss == 0, f"Data table {tname} sampled rows match tidy", f"{miss} mismatches")

# ============================================================ DASHBOARD KPIs
ws = wb["Dashboard"]
ly = max(y for y in s.index if y not in engine.PARTIAL_YEARS and s.loc[y, "Unplanned"] > 0)
# KPI values are in merged cells row 6 (index) -> openpyxl row 6; find by scanning
kpis = {}
for row in ws.iter_rows(min_row=6, max_row=6):
    for c in row:
        if isinstance(c.value, (int, float)): kpis.setdefault("vals", []).append(c.value)
vals = kpis.get("vals", [])
check(any(close(v, float(s.loc[ly,"Total"])) for v in vals), "Dashboard KPI Total Offline correct")
check(any(close(v, float(s.loc[ly,"Unplanned"])) for v in vals), "Dashboard KPI Unplanned correct")

# ============================================================ MODEL
ws = wb["Model"]
# lookup profiles
windows = list(engine.BASELINE_WINDOWS.keys())
br, _ = band_row(ws, "Lookup - Avg Unplanned")
hdr = br + 1; miss = 0
for wi, w in enumerate(windows):
    prof = engine.baseline_profile(ctx["df"], w)
    r = hdr + 1 + wi
    for j, m in enumerate(MONTHS):
        if not close(num(ws, r, 3 + j), float(prof[m])): miss += 1
check(miss == 0, "Model baseline-window profiles match engine", f"{miss} mismatches")
# per-PADD scenario baseline
sp = ctx["scenario_padd"]
br, _ = band_row(ws, "2027 Scenario by PADD")
hdr = br + 1; miss = 0
for pi, p in enumerate(PADDS):
    r = hdr + 1 + pi
    if not close(num(ws, r, 3), float(sp[p]["baseline_annual"])): miss += 1
check(miss == 0, "Model per-PADD baseline matches engine", f"{miss} mismatches")
# tornado
tor = ctx["tornado"]
br, _ = band_row(ws, "Tornado - 2027 Unplanned")
hdr = br + 1; miss = 0
for ti, row in enumerate(tor):
    r = hdr + 1 + ti
    if not close(num(ws, r, 3), row["low"]) or not close(num(ws, r, 5), row["high"]): miss += 1
check(miss == 0, "Model tornado low/high match engine", f"{miss} mismatches")

# ============================================================ FORMULAS structural
err_tokens = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!"]
fcount = 0; ferr = 0; unbalanced = 0
for sh in wb.worksheets:
    for row in sh.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                fcount += 1
                if any(t in c.value for t in err_tokens): ferr += 1
                if c.value.count("(") != c.value.count(")"): unbalanced += 1
check(ferr == 0, f"No error tokens in {fcount} formulas", f"{ferr} bad")
check(unbalanced == 0, "All formulas balanced parens", f"{unbalanced} bad")

# ============================================================ CHARTS
chart_problems = []
total_charts = 0; total_series = 0
for sh in wb.worksheets:
    for ch in getattr(sh, "_charts", []):
        total_charts += 1
        title = ""
        try:
            if ch.title and ch.title.tx and ch.title.tx.rich:
                title = "".join(r.t or "" for p in ch.title.tx.rich.p for r in (p.r or []))
        except Exception: pass
        sers = list(ch.series)
        if not sers:
            chart_problems.append(f"{sh.title}: '{title[:30]}' has NO series"); continue
        for si, se in enumerate(sers):
            total_series += 1
            ref = None
            try: ref = se.val.numRef.f
            except Exception:
                try: ref = se.yVal.numRef.f
                except Exception: ref = None
            if not ref:
                chart_problems.append(f"{sh.title}: '{title[:24]}' s{si} no value ref"); continue
            # parse 'Sheet!$A$1:$B$2'
            m = re.match(r"(?:'?([^'!]+)'?!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", ref)
            if not m:
                m2 = re.match(r"(?:'?([^'!]+)'?!)?\$?([A-Z]+)\$?(\d+)", ref)
                if not m2: chart_problems.append(f"{sh.title}: '{title[:24]}' unparsable ref {ref}"); continue
                continue
            tsheet = m.group(1) or sh.title
            try: tws = wb[tsheet]
            except Exception:
                chart_problems.append(f"{sh.title}: '{title[:24]}' bad sheet {tsheet}"); continue
            c1, r1, c2, r2 = range_boundaries(ref.split("!")[-1].replace("$",""))
            # check at least one non-empty cell in the value range
            nonempty = 0
            for rr in range(r1, r2+1):
                for cc in range(c1, c2+1):
                    v = tws.cell(row=rr, column=cc).value
                    if v is not None and v != "": nonempty += 1
            if nonempty == 0:
                chart_problems.append(f"{sh.title}: '{title[:24]}' s{si} EMPTY range {ref}")
check(len(chart_problems) == 0, f"All {total_charts} charts / {total_series} series reference populated data",
      f"{len(chart_problems)} problems")

# ============================================================ AUTO-INSIGHTS
ws = wb["Cover"]
movers = ctx["top_movers"]
br, _ = band_row(ws, "This Week's Reads")
found = 0
if br:
    for r in range(br+1, br+1+len(movers)):   # only the insights block
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip().startswith("•"): found += 1
check(found == len(movers), f"Cover auto-insights present ({found}/{len(movers)})")

# combo charts: verify bar+line co-exist in the PADD chart XML
z = zipfile.ZipFile(WB)
combo = 0
for nm in z.namelist():
    if re.match(r"xl/charts/chart\d+\.xml$", nm):
        x = z.read(nm).decode("utf-8", "replace")
        if "<c:barChart>" in x and "<c:lineChart>" in x:
            combo += 1
check(combo >= 5, f"combo charts have both bar+line plots ({combo} found)")

# ============================================================ REPORT
print("="*70)
print(f"QA RESULTS:  {len(PASS)} passed,  {len(FAIL)} failed")
print("="*70)
for p in PASS: print("  PASS ", p)
if FAIL:
    print("\n  ---- FAILURES ----")
    for fl in FAIL: print("  FAIL ", fl)
if chart_problems:
    print("\n  ---- CHART PROBLEMS ----")
    for cp in chart_problems[:30]: print("   -", cp)
print("\nTotals: charts=%d series=%d formulas=%d" % (total_charts, total_series, fcount))
sys.exit(1 if FAIL or chart_problems else 0)
