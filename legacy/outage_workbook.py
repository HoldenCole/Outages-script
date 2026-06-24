#!/usr/bin/env python3
"""Excel workbook generator for the refinery outage analysis (KBD).

Targeted workbook builder for the business comparison years 2025, 2026, and
2027. It produces:
  * Summary sheet with exact pairwise comparisons:
      - 2025 vs 2026
      - 2025 vs 2027
      - 2026 vs 2027
    for Plan + Unplanned and Planned
  * Unplanned section only compares years with actual unplanned data:
      - 2025 vs 2026
  * PADD 1..5 sheets with the same exact pairwise blocks
  * A compact scenario sheet for 2027 planned + historical unplanned bands
"""
import argparse
import sys

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from outage_monthly import (
    COLUMN_MAP,
    MONTHS,
    REQUIRED,
    _classify_type,
    _norm,
    combined_over_types,
    monthly_matrix,
    monthly_row,
    period_summaries,
)

REGIONS = ["Total US", "PADD 1", "PADD 2", "PADD 3", "PADD 4", "PADD 5"]
PAIR_YEARS_ALL = [(2025, 2026), (2025, 2027), (2026, 2027)]
PAIR_YEARS_UNPLANNED = [(2025, 2026)]
TARGET_YEARS = [2025, 2026, 2027]
SCENARIO_YEAR = 2027
SCENARIO_HIST_YEARS = [2025, 2026]

HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
ANCHOR_FILL = PatternFill("solid", fgColor="E8EEF5")
TITLE_FILL = PatternFill("solid", fgColor="14181D")
POS_FILL = PatternFill("solid", fgColor="C9EBD0")
NEG_FILL = PatternFill("solid", fgColor="F6C9C6")

HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="E8A13A", bold=True, size=12)
ANCHOR_FONT = Font(bold=True, color="14181D")
LABEL_FONT = Font(color="14181D")
NUM_FONT = Font(name="Consolas", color="14181D")

THIN = Side(style="thin", color="C9D2DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

KBD_FMT = "#,##0"
PCT_FMT = "0%"

FIRST_MONTH_COL = 2
LAST_MONTH_COL = FIRST_MONTH_COL + 11
GAP_COL = LAST_MONTH_COL + 1
TOTAL_COL = GAP_COL + 1
H1_COL = TOTAL_COL + 1
H2_COL = H1_COL + 1
LAST_COL = H2_COL


def load_any(path):
    lower = path.lower()
    try:
        if lower.endswith(".csv"):
            raw = pd.read_csv(path)
        else:
            raw = pd.read_excel(path)
    except FileNotFoundError:
        raise SystemExit(f"ERROR: could not read input file '{path}': file not found.")
    except Exception as e:
        raise SystemExit(f"ERROR: could not read input file '{path}': {e}")

    actual = {_norm(c): c for c in raw.columns}
    rename, found, missing = {}, {}, []
    for logical, header in COLUMN_MAP.items():
        key = _norm(header)
        if key in actual:
            rename[actual[key]] = logical
            found[logical] = actual[key]
        else:
            missing.append((logical, header))

    df = raw.rename(columns=rename)
    df = df[[c for c in df.columns if c in COLUMN_MAP]].copy()

    missing_required = [m for m in missing if m[0] in REQUIRED]
    if missing_required:
        lines = "\n".join(f" - logical '{lg}' expected header '{hd}'" for lg, hd in missing_required)
        have = "\n".join(f" - {c}" for c in raw.columns)
        raise SystemExit(
            "ERROR: required columns for the monthly matrix not found.\n"
            + lines
            + "\n\nHeaders present in the file:\n"
            + have
        )

    for c in ("kbd", "year", "month", "pct_month", "pct_month_cal", "total_days", "month_days", "offline_cap"):
        if c in df:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    for c in ("start_date", "end_date", "month_date"):
        if c in df:
            df[c] = pd.to_datetime(df[c], errors="coerce")

    if ("month" not in df or df["month"].isna().all()):
        if "month_date" in df and df["month_date"].notna().any():
            df["month"] = df["month_date"].dt.month
        elif "start_date" in df:
            df["month"] = df["start_date"].dt.month

    if "padd" in df:
        raw_padd = df["padd"].copy()
        parsed = pd.to_numeric(raw_padd, errors="coerce")
        extracted = raw_padd.astype(str).str.extract(r"(\d)", expand=False)
        extracted = pd.to_numeric(extracted, errors="coerce")
        df["padd"] = parsed.fillna(extracted)

    if "outage_type" in df:
        df["type_bucket"] = df["outage_type"].map(_classify_type)
    else:
        df["type_bucket"] = "unknown"

    return df, found, [m for m in missing if m[0] not in REQUIRED]


def build_full_matrix(df):
    base = monthly_matrix(df, value="kbd", convention="as_is")
    allt = combined_over_types(base)
    full = pd.concat([base, allt], ignore_index=True)
    full["region"] = full["region"].astype(str).str.strip()
    return full


def years_present(matrix):
    return sorted({int(y) for y in matrix["year"].dropna().unique()})


def row_for(matrix, region, year, bucket):
    region = str(region).strip()
    vals = monthly_row(matrix, region, year, bucket)
    return vals


def pct_vs(base, comp):
    out = []
    for b, c in zip(base, comp):
        if b is None or c is None or (isinstance(b, float) and np.isnan(b)) or b == 0:
            out.append(None)
        else:
            out.append((c - b) / b)
    return out


def band_percentiles(matrix, region, years):
    series_by_month = {m: [] for m in range(1, 13)}
    for y in years:
        vals = monthly_row(matrix, region, y, "unplanned")
        for i, v in enumerate(vals):
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                series_by_month[i + 1].append(v)
    low, base, high = [], [], []
    for m in range(1, 13):
        xs = series_by_month[m]
        if xs:
            low.append(float(np.percentile(xs, 25)))
            base.append(float(np.percentile(xs, 50)))
            high.append(float(np.percentile(xs, 90)))
        else:
            low.append(0.0)
            base.append(0.0)
            high.append(0.0)
    return low, base, high


class SheetWriter:
    def __init__(self, ws):
        self.ws = ws
        self.r = 1
        self.month_data_rows = []

    def _style_label(self, cell, anchor=False):
        cell.font = ANCHOR_FONT if anchor else LABEL_FONT
        cell.alignment = LEFT
        cell.border = BORDER
        if anchor:
            cell.fill = ANCHOR_FILL

    def _style_num(self, cell, fmt, anchor=False):
        cell.font = NUM_FONT if not anchor else Font(name="Consolas", bold=True, color="14181D")
        cell.alignment = RIGHT
        cell.number_format = fmt
        cell.border = BORDER
        if anchor:
            cell.fill = ANCHOR_FILL

    def blank(self, n=1):
        self.r += n

    def title(self, text):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=LAST_COL)
        c = ws.cell(row=self.r, column=1, value=text)
        c.font = TITLE_FONT
        c.fill = TITLE_FILL
        c.alignment = LEFT
        self.r += 1

    def header(self):
        ws = self.ws
        ws.cell(row=self.r, column=1, value="")
        labels = MONTHS + ["", "Total", "1H Avg", "2H Avg"]
        for j, lab in enumerate(labels):
            col = FIRST_MONTH_COL + j
            c = ws.cell(row=self.r, column=col, value=lab)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = BORDER
        lc = ws.cell(row=self.r, column=1)
        lc.fill = HEADER_FILL
        lc.border = BORDER
        self.r += 1

    def kbd_row(self, label, vals, anchor=False, fmt=KBD_FMT, track=True):
        ws = self.ws
        ridx = self.r
        lc = ws.cell(row=ridx, column=1, value=label)
        self._style_label(lc, anchor=anchor)
        for i in range(12):
            col = FIRST_MONTH_COL + i
            v = vals[i]
            cell = ws.cell(row=ridx, column=col, value=(None if v is None else round(v, 1)))
            self._style_num(cell, fmt, anchor=anchor)
        s = period_summaries([0 if v is None else v for v in vals])
        tot = sum(v for v in vals if v is not None)
        for col, val in ((TOTAL_COL, tot), (H1_COL, s["h1_avg"]), (H2_COL, s["h2_avg"])):
            cell = ws.cell(row=ridx, column=col, value=round(val, 1))
            self._style_num(cell, fmt, anchor=anchor)
        if track:
            self.month_data_rows.append((ridx, label))
        self.r += 1
        return ridx

    def pct_row(self, label, base, comp):
        ws = self.ws
        ridx = self.r
        lc = ws.cell(row=ridx, column=1, value=label)
        lc.font = Font(italic=True, color="5B6776")
        lc.alignment = LEFT
        lc.border = BORDER
        pcts = pct_vs(base, comp)
        for i in range(12):
            col = FIRST_MONTH_COL + i
            v = pcts[i]
            cell = ws.cell(row=ridx, column=col, value=(None if v is None else v))
            cell.alignment = RIGHT
            cell.number_format = PCT_FMT
            cell.border = BORDER
            cell.font = Font(name="Consolas", size=10, color="14181D")
            if v is not None:
                cell.fill = POS_FILL if v >= 0 else NEG_FILL
        bs = period_summaries([0 if v is None else v for v in base])
        cs = period_summaries([0 if v is None else v for v in comp])
        base_tot = sum(v for v in base if v is not None)
        comp_tot = sum(v for v in comp if v is not None)
        trailing = pct_vs([base_tot, bs["h1_avg"], bs["h2_avg"]], [comp_tot, cs["h1_avg"], cs["h2_avg"]])
        for col, v in zip((TOTAL_COL, H1_COL, H2_COL), trailing):
            cell = ws.cell(row=ridx, column=col, value=(None if v is None else v))
            cell.alignment = RIGHT
            cell.number_format = PCT_FMT
            cell.border = BORDER
            cell.font = Font(name="Consolas", size=10, color="14181D")
            if v is not None:
                cell.fill = POS_FILL if v >= 0 else NEG_FILL
        self.r += 1
        return ridx


def label_for_bucket(bucket):
    if bucket == "planned":
        return "Plan"
    if bucket == "unplanned":
        return "Unplanned Outages"
    return "Outages"


def write_pair_block(sw, matrix, region, first_year, second_year, bucket, title):
    sw.title(title)
    sw.header()
    first_vals = row_for(matrix, region, first_year, bucket)
    second_vals = row_for(matrix, region, second_year, bucket)
    sw.kbd_row(f"{first_year} {label_for_bucket(bucket)}", first_vals, anchor=True)
    sw.kbd_row(f"{second_year} {label_for_bucket(bucket)}", second_vals)
    sw.pct_row(f"{second_year} vs {first_year}", first_vals, second_vals)
    sw.blank()


def write_region_tables(sw, matrix, region):
    sw.title(f"{region} targeted comparisons (KBD)")
    sw.blank()

    for a, b in PAIR_YEARS_ALL:
        write_pair_block(sw, matrix, region, a, b, "all", f"Plan + Unplanned — {a} vs {b}")
    sw.blank()

    for a, b in PAIR_YEARS_ALL:
        write_pair_block(sw, matrix, region, a, b, "planned", f"Planned — {a} vs {b}")
    sw.blank()

    for a, b in PAIR_YEARS_UNPLANNED:
        write_pair_block(sw, matrix, region, a, b, "unplanned", f"Unplanned — {a} vs {b}")
    sw.blank(2)


def write_scenario_region(sw, matrix, region, scenario_year=SCENARIO_YEAR, hist_years=SCENARIO_HIST_YEARS):
    sw.title(f"{region} {scenario_year} Scenario — Planned + historical-unplanned bands (KBD)")
    sw.header()
    planned = row_for(matrix, region, scenario_year, "planned")
    low, base, high = band_percentiles(matrix, region, hist_years)
    add = lambda a, b: [(None if (a[i] is None and b[i] is None) else (a[i] or 0) + (b[i] or 0)) for i in range(12)]
    sw.kbd_row(f"{scenario_year} Planned", planned, anchor=True)
    sw.kbd_row(f"{scenario_year} Planned + Low (P25)", add(planned, low))
    sw.kbd_row(f"{scenario_year} Planned + Base (P50)", add(planned, base))
    sw.kbd_row(f"{scenario_year} Planned + High (P90)", add(planned, high))
    sw.blank()
    for y in sorted(hist_years, reverse=True):
        sw.kbd_row(f"{y} Unplanned (actual)", row_for(matrix, region, y, "unplanned"), track=False)
    sw.blank(2)


def _header_row_index(ws):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=FIRST_MONTH_COL).value == "Jan":
            return row
    return 2


def add_combo_chart(ws, title, bar_row, line_row, anchor_cell):
    bar = BarChart()
    bar.type = "col"
    bar.title = title
    bar.y_axis.title = "KBD"
    bar.x_axis.title = "Month"
    bar.height = 7.5
    bar.width = 18
    hdr = _header_row_index(ws)
    cats = Reference(ws, min_col=FIRST_MONTH_COL, max_col=LAST_MONTH_COL, min_row=hdr, max_row=hdr)
    bdata = Reference(ws, min_col=FIRST_MONTH_COL, max_col=LAST_MONTH_COL, min_row=bar_row, max_row=bar_row)
    bar.add_data(bdata, from_rows=True)
    bar.set_categories(cats)
    line = LineChart()
    ldata = Reference(ws, min_col=FIRST_MONTH_COL, max_col=LAST_MONTH_COL, min_row=line_row, max_row=line_row)
    line.add_data(ldata, from_rows=True)
    line.set_categories(cats)
    bar += line
    ws.add_chart(bar, anchor_cell)


def finalize_sheet(ws, freeze=True):
    ws.sheet_view.showGridLines = False
    if freeze:
        hdr = _header_row_index(ws)
        ws.freeze_panes = ws.cell(row=hdr + 1, column=FIRST_MONTH_COL).coordinate
    ws.column_dimensions["A"].width = 28
    for col in range(FIRST_MONTH_COL, LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 7.5
    ws.column_dimensions[get_column_letter(GAP_COL)].width = 2
    ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 8.5
    ws.column_dimensions[get_column_letter(H1_COL)].width = 8.5
    ws.column_dimensions[get_column_letter(H2_COL)].width = 8.5


def debug_summary(df, matrix, years):
    print("\nDEBUG SUMMARY")
    print("=" * 78)
    print(f"rows loaded: {len(df):,}")
    print(f"years present: {', '.join(str(y) for y in years)}")
    if "type_bucket" in df:
        print(f"type buckets: {df['type_bucket'].value_counts().to_dict()}")
    if "padd" in df:
        non_null_padds = sorted(df["padd"].dropna().unique().tolist())
        print(f"raw padds present: {non_null_padds}")
        print(f"raw padd counts: {df['padd'].value_counts(dropna=False).to_dict()}")
    print(f"matrix regions: {sorted(matrix['region'].dropna().unique().tolist())}")
    for rg in REGIONS:
        print(f"\n[{rg}]")
        for yr in TARGET_YEARS:
            print(f"  {yr} all       : {row_for(matrix, rg, yr, 'all')}")
            print(f"  {yr} planned   : {row_for(matrix, rg, yr, 'planned')}")
            print(f"  {yr} unplanned : {row_for(matrix, rg, yr, 'unplanned')}")
    print("=" * 78)


def build_workbook(df, out_path):
    matrix = build_full_matrix(df)
    all_years = years_present(matrix)
    missing = [y for y in TARGET_YEARS if y not in all_years]
    if missing:
        raise SystemExit(f"ERROR: required comparison years missing from dataset: {missing}")

    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    sw = SheetWriter(ws_sum)
    write_region_tables(sw, matrix, "Total US")
    finalize_sheet(ws_sum)

    for region in REGIONS:
        if region == "Total US":
            continue
        ws = wb.create_sheet(title=region)
        sw = SheetWriter(ws)
        write_region_tables(sw, matrix, region)
        finalize_sheet(ws)

        row_2025 = next((r for r, lbl in sw.month_data_rows if lbl == "2025 Outages"), None)
        row_2026 = next((r for r, lbl in sw.month_data_rows if lbl == "2026 Outages"), None)
        if row_2025 and row_2026:
            add_combo_chart(ws, f"{region}: 2025 vs 2026 plan + unplanned", row_2025, row_2026, ws.cell(row=ws.max_row + 2, column=1).coordinate)

    ws_sc = wb.create_sheet(title="2027 Scenario")
    sw = SheetWriter(ws_sc)
    for region in REGIONS:
        write_scenario_region(sw, matrix, region)
    finalize_sheet(ws_sc)

    wb.save(out_path)
    return {
        "years": all_years,
        "sheets": [ws.title for ws in wb.worksheets],
        "out": out_path,
        "matrix": matrix,
    }


def _selftest():
    pv = pct_vs([100, 0, 0], [110, 0, 50])
    assert round(pv[0], 10) == 0.1
    assert pv[1] is None
    assert pv[2] is None
    row = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 110, 120]
    s = period_summaries(row)
    assert sum(row) == 780
    assert s["h1_avg"] == round((10 + 20 + 30 + 40 + 50 + 60) / 6, 1)
    assert s["h2_avg"] == round((70 + 80 + 90 + 100 + 110 + 120) / 6, 1)
    print("selftest OK")


def main():
    ap = argparse.ArgumentParser(description="Refinery outage Excel workbook generator (KBD)")
    ap.add_argument("excel", nargs="?", help="path to the outage .xlsx or .csv export")
    ap.add_argument("--out", default="outage_workbook.xlsx", help="output .xlsx path")
    ap.add_argument("--selftest", action="store_true", help="run synthetic-data sanity checks and exit")
    ap.add_argument("--debug-summary", action="store_true", help="print detected rows/years/PADD breakdowns before writing the workbook")
    args = ap.parse_args()

    if args.selftest:
        _selftest()
        return
    if not args.excel:
        ap.error("the 'excel' input path is required (or pass --selftest)")

    print(f"Reading {args.excel} ...")
    df, found, missing_optional = load_any(args.excel)
    print(f"  {len(df):,} rows loaded.")
    if found:
        print("  matched columns:")
        for lg, hd in found.items():
            print(f"    - {lg}: {hd}")
    if missing_optional:
        print("  optional columns not found:")
        for lg, hd in missing_optional:
            print(f"    - {lg} (expected '{hd}')")

    info = build_workbook(df, args.out)
    print(f"  years present: {', '.join(str(y) for y in info['years'])}")
    print(f"  sheets: {', '.join(info['sheets'])}")

    if args.debug_summary:
        debug_summary(df, info["matrix"], info["years"])

    print(f"\nDone -> {info['out']}")


if __name__ == "__main__":
    try:
        main()
    except SystemExit as e:
        if isinstance(e.code, str):
            print(e.code)
            sys.exit(1)
        raise
